# Fabric notebook source

# METADATA ********************

# META {
# META   "kernel_info": {
# META     "name": "synapse_pyspark"
# META   },
# META   "dependencies": {
# META     "lakehouse": {
# META       "default_lakehouse": "086407fd-b321-4254-b11b-0f2c1677636c",
# META       "default_lakehouse_name": "LH_Param_Demo",
# META       "default_lakehouse_workspace_id": "8b9e9590-f406-49ee-a8fa-43321bf8d391",
# META       "known_lakehouses": [
# META         {
# META           "id": "086407fd-b321-4254-b11b-0f2c1677636c"
# META         }
# META       ]
# META     }
# META   }
# META }

# MARKDOWN ********************

# # 📊 Bronze Layer Data Profiling Notebook
# ### Source-agnostic | Auto-discovers new tables | No code changes needed when Bronze grows
# 
# **What this notebook does**
# - Scans every table in the schema(s) you list below (e.g. `bronze`)
# - Auto-detects the **source system** (Acumatica / Oracle IFS / SharePoint / OneDrive / etc.) from the table naming convention
# - Profiles each table: row count, size, primary key candidate, incremental/date column, CDC availability, min/max dates, estimated daily growth
# - Exports everything to a single **Excel file** in the Lakehouse **Files** section (not a Delta table) so you can download it from the Fabric UI
# 
# **When you add new Bronze tables:** just re-run the notebook. If a table's source can't be auto-classified, it will show up clearly labeled `Unknown / Unclassified` so you know to add a mapping entry — the run never breaks.
# 
# **The only cell you should normally touch is the CONFIGURATION cell below.**


# CELL ********************

# =========================================================
# IMPORTS
# =========================================================
import os
import re
from datetime import datetime, date

import pandas as pd
from pyspark.sql import functions as F
from pyspark.sql.types import DateType, TimestampType

from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl import Workbook


# METADATA ********************

# META {
# META   "language": "python",
# META   "language_group": "synapse_pyspark"
# META }

# CELL ********************

# =========================================================
# CONFIGURATION  ⚙️  <-- Edit only this cell when your environment changes
# =========================================================

# Schemas to scan for tables (add "silver", "gold" etc. later if you want them profiled too)
SCHEMAS_TO_SCAN = ["bronze"]

# Leave as None to auto-use the lakehouse currently attached to this notebook
DATABASE_NAME = None

# Output Excel file
OUTPUT_FILE_NAME = f"Data_Profiling_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

# Files (NOT Tables) area of the default lakehouse -- this is just a landing spot for the
# .xlsx so you can download it from the Fabric UI. It is not saved as a lakehouse table.
OUTPUT_DIR = "/lakehouse/default/Files/data_profiling_output"

# ---- SOURCE SYSTEM DETECTION RULES ----------------------------------------------------
# Table naming convention observed: bronze_<sourcecode>_<entity_name>
# Extend this dict (not the code) whenever a brand-new source system is onboarded.
SOURCE_SYSTEM_PREFIX_MAP = {
    "acu":        "Acumatica",
    "ifs":        "Oracle IFS",
    "sp":         "SharePoint Online List",
    "sharepoint": "SharePoint Online List",
    "od":         "OneDrive",
    "onedrive":   "OneDrive",
}

# ---- OPTIONAL MANUAL METADATA OVERRIDES ------------------------------------------------
# Purpose / Entity Type / Refresh Frequency / business Foreign Keys can't be derived
# automatically. Add entries here keyed by EXACT table name to enrich the report.
# Leaving a table out is fine -- it just shows up with defaults ("Not Set", blank, etc.)
MANUAL_METADATA = {
    # "bronze_acu_newcustomersolid": {
    #     "Entity Type": "Dimension",
    #     "Purpose": "Customer master data used for AR reporting",
    #     "Refresh Frequency": "Daily",
    #     "Foreign Keys": "CustomerID -> bronze_acu_ar_customers.CustomerID",
    # },
}

# ---- PERFORMANCE CONTROLS --------------------------------------------------------------
ENABLE_PK_DETECTION   = True        # set False to skip PK heuristic entirely (fastest)
SAMPLE_ROW_THRESHOLD  = 5_000_000   # tables bigger than this get sampled for PK/date checks
SAMPLE_FRACTION       = 0.02        # 2% sample used for big tables


# METADATA ********************

# META {
# META   "language": "python",
# META   "language_group": "synapse_pyspark"
# META }

# CELL ********************

# =========================================================
# SOURCE SYSTEM DETECTION
# =========================================================
def detect_source_system(table_name: str) -> str:
    """Infer source system from bronze_<code>_<entity> naming convention,
    falling back to a substring match anywhere in the table name."""
    parts = table_name.lower().split("_")
    if len(parts) >= 2:
        candidate = parts[1]
        if candidate in SOURCE_SYSTEM_PREFIX_MAP:
            return SOURCE_SYSTEM_PREFIX_MAP[candidate]

    lname = table_name.lower()
    for prefix, source in sorted(SOURCE_SYSTEM_PREFIX_MAP.items(), key=lambda x: -len(x[0])):
        if prefix in lname:
            return source

    return "Unknown / Unclassified (add to SOURCE_SYSTEM_PREFIX_MAP)"


# METADATA ********************

# META {
# META   "language": "python",
# META   "language_group": "synapse_pyspark"
# META }

# CELL ********************

# =========================================================
# TABLE DISCOVERY
# =========================================================
def get_all_tables(schema_name: str):
    tables = []
    try:
        for t in spark.catalog.listTables(schema_name):
            if not t.isTemporary:
                tables.append(t.name)
    except Exception as e:
        print(f"[WARN] Could not list tables for schema '{schema_name}': {e}")
    return tables


# METADATA ********************

# META {
# META   "language": "python",
# META   "language_group": "synapse_pyspark"
# META }

# CELL ********************

# =========================================================
# CORE PROFILING LOGIC
# =========================================================
def profile_table(schema_name: str, table_name: str) -> dict:
    full_name = f"{schema_name}.{table_name}"

    row = {
        "Source System": detect_source_system(table_name),
        "Database": DATABASE_NAME if DATABASE_NAME else spark.catalog.currentDatabase(),
        "Schema": schema_name,
        "Table/View Name": table_name,
        "Entity Type": MANUAL_METADATA.get(table_name, {}).get("Entity Type", "Unclassified"),
        "Purpose": MANUAL_METADATA.get(table_name, {}).get("Purpose", ""),
        "Primary Key": "",
        "Foreign Keys": MANUAL_METADATA.get(table_name, {}).get("Foreign Keys", ""),
        "Record Count": None,
        "Daily Growth (Est.)": None,
        "Estimated Size (GB)": None,
        "Refresh Frequency": MANUAL_METADATA.get(table_name, {}).get("Refresh Frequency", "Not Set"),
        "Incremental Column": "",
        "CDC Available": "No",
        "MAX Date": "",
        "Historical Data Available (Earliest Date)": "",
        "Notes / Issues": "",
    }

    try:
        df = spark.table(full_name)

        # ---- Record count ----
        record_count = df.count()
        row["Record Count"] = record_count

        # ---- Size on disk ----
        try:
            detail = spark.sql(f"DESCRIBE DETAIL {full_name}").collect()[0]
            size_bytes = detail["sizeInBytes"] or 0
            row["Estimated Size (GB)"] = round(size_bytes / (1024 ** 3), 4)
        except Exception:
            pass

        # ---- CDC availability (Delta Change Data Feed) ----
        try:
            props = {r["key"]: r["value"] for r in spark.sql(f"SHOW TBLPROPERTIES {full_name}").collect()}
            row["CDC Available"] = "Yes" if props.get("delta.enableChangeDataFeed", "false").lower() == "true" else "No"
        except Exception:
            pass

        # ---- Sampling for expensive checks on very large tables ----
        use_sample = record_count > SAMPLE_ROW_THRESHOLD
        scan_df = df.sample(fraction=SAMPLE_FRACTION, seed=42) if use_sample else df
        if use_sample:
            row["Notes / Issues"] += (
                f"Large table ({record_count:,} rows) - PK/date checks computed on "
                f"a {int(SAMPLE_FRACTION*100)}% sample. "
            )

        # ---- Primary key candidate ----
        if ENABLE_PK_DETECTION:
            pk_candidates = [c for c in df.columns if re.search(r"(^id$|_id$|id$)", c, re.IGNORECASE)]
            if not pk_candidates:
                pk_candidates = [df.columns[0]]

            best_pk = None
            basis_count = scan_df.count() if use_sample else record_count
            for col in pk_candidates[:5]:
                try:
                    nulls = scan_df.filter(F.col(col).isNull()).count()
                    distincts = scan_df.select(F.approx_count_distinct(col)).collect()[0][0]
                    if basis_count > 0 and nulls == 0 and distincts >= basis_count * 0.995:
                        best_pk = col
                        break
                except Exception:
                    continue
            row["Primary Key"] = best_pk if best_pk else "Not Detected"

        # ---- Incremental / date column detection ----
        date_pattern = re.compile(r"(date|_dt$|^dt$|time|modified|updated|created|timestamp)", re.IGNORECASE)
        date_cols = [
            f.name for f in df.schema.fields
            if isinstance(f.dataType, (DateType, TimestampType)) or date_pattern.search(f.name)
        ]

        if date_cols:
            preferred = [c for c in date_cols if re.search(r"(modified|updated|change)", c, re.IGNORECASE)]
            incremental_col = preferred[0] if preferred else date_cols[0]
            row["Incremental Column"] = incremental_col

            try:
                stats = scan_df.select(
                    F.min(F.col(incremental_col)).alias("min_d"),
                    F.max(F.col(incremental_col)).alias("max_d"),
                ).collect()[0]

                row["MAX Date"] = str(stats["max_d"]) if stats["max_d"] is not None else ""
                row["Historical Data Available (Earliest Date)"] = (
                    str(stats["min_d"]) if stats["min_d"] is not None else ""
                )

                if stats["min_d"] and stats["max_d"]:
                    try:
                        d_min = stats["min_d"] if isinstance(stats["min_d"], (date, datetime)) else datetime.fromisoformat(str(stats["min_d"]))
                        d_max = stats["max_d"] if isinstance(stats["max_d"], (date, datetime)) else datetime.fromisoformat(str(stats["max_d"]))
                        days_span = max((d_max - d_min).days, 1)
                        row["Daily Growth (Est.)"] = round(record_count / days_span, 2)
                    except Exception:
                        row["Daily Growth (Est.)"] = "N/A"
            except Exception as e:
                row["Notes / Issues"] += f"Date profiling failed: {e}. "
        else:
            row["Incremental Column"] = "None Detected"

    except Exception as e:
        row["Notes / Issues"] += f"Profiling error: {e}"

    return row


# METADATA ********************

# META {
# META   "language": "python",
# META   "language_group": "synapse_pyspark"
# META }

# CELL ********************

# =========================================================
# RUN THE PROFILING ACROSS ALL SCHEMAS/TABLES
# =========================================================
all_results = []

for schema in SCHEMAS_TO_SCAN:
    print(f"Scanning schema: {schema}")
    tables = get_all_tables(schema)
    print(f"  Found {len(tables)} tables")
    for i, t in enumerate(tables, 1):
        print(f"  ({i}/{len(tables)}) Profiling {schema}.{t} ...")
        all_results.append(profile_table(schema, t))

print(f"\nDone. Profiled {len(all_results)} tables total.")


# METADATA ********************

# META {
# META   "language": "python",
# META   "language_group": "synapse_pyspark"
# META }

# CELL ********************

# =========================================================
# BUILD THE FINAL DATAFRAME
# =========================================================
COLUMN_ORDER = [
    "Source System", "Database", "Schema", "Table/View Name", "Entity Type", "Purpose",
    "Primary Key", "Foreign Keys", "Record Count", "Daily Growth (Est.)",
    "Estimated Size (GB)", "Refresh Frequency", "Incremental Column", "CDC Available",
    "MAX Date", "Historical Data Available (Earliest Date)", "Notes / Issues",
]

pdf = pd.DataFrame(all_results)[COLUMN_ORDER]
pdf = pdf.sort_values(["Source System", "Schema", "Table/View Name"]).reset_index(drop=True)

display(pdf)


# METADATA ********************

# META {
# META   "language": "python",
# META   "language_group": "synapse_pyspark"
# META }

# CELL ********************

# =========================================================
# EXPORT TO EXCEL (formatted) -- landed in Lakehouse Files, NOT saved as a table
# =========================================================
os.makedirs(OUTPUT_DIR, exist_ok=True)
output_path = os.path.join(OUTPUT_DIR, OUTPUT_FILE_NAME)

wb = Workbook()
ws = wb.active
ws.title = "Data Profiling"

for r in dataframe_to_rows(pdf, index=False, header=True):
    ws.append(r)

header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions

for col_cells in ws.columns:
    length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
    col_letter = get_column_letter(col_cells[0].column)
    ws.column_dimensions[col_letter].width = min(max(length + 3, 12), 45)

wb.save(output_path)
print(f"Excel report saved to: {output_path}")
print("Go to your Lakehouse's 'Files' pane -> data_profiling_output -> right-click the file -> Download.")


# METADATA ********************

# META {
# META   "language": "python",
# META   "language_group": "synapse_pyspark"
# META }
